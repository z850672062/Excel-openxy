# -*- coding: utf-8 -*-
import socket
import os
import time
import wmi
import win32com.client
from openpyxl import load_workbook

filename = "电脑信息记录.xlsm"
c = wmi.WMI ()
# 输入使用人
name = input("请输入用户姓名,以回车键结束。")
# print(name)
print('\n'+"运行中，请稍等。")
# SN
strComputer = "."
objWMIService = win32com.client.Dispatch("WbemScripting.SWbemLocator")
objSWbemServices = objWMIService.ConnectServer(strComputer,"root\cimv2")
colItems = objSWbemServices.ExecQuery("Select * From Win32_BIOS")
SerialNumber=colItems[0].SerialNumber
# print (SerialNumber)

# 当前时间
localtime =time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())
# print(localtime)



# 计算机名字
hostname = socket.gethostname()
# 当前内网ip
ip = socket.gethostbyname(hostname)

# print("hostname:",hostname)
# print("ip:",ip)



# 获取网卡信息
macs = []
for n in  c.Win32_NetworkAdapter():
    mactmp = n.MACAddress
    if mactmp and len(mactmp.strip()) > 5:
        tmpmsg = {}
        tmpmsg[n.Name] = n.MACAddress
        macs.append(tmpmsg)
        print(tmpmsg)

# #
# len1 = len(macs)
# print(len1)
# # for i in macs:
# #     print(i)
# # 卡1
Eth_GbE = ""
Wireless = ""
mac1 = []
if len(macs) >0:
    for i in range(len((macs))):
        a =macs[i]
        test3 = min(a, key=lambda x: a[x])
        # 判断键中是否有Ethernet字体 有返回true
        if str.find(test3, "Ethernet") != -1:
            print("有线MAC:", a.get("%s" % test3))
            mac1.append(a.get("%s" % test3))
            Eth_GbE = "Eth_GbE:"+ a.get("%s" % test3)
        elif str.find(test3, "GbE") != -1:
            print("有线MAC:", a.get("%s" % test3))
            mac1.append(a.get("%s" % test3))
            Eth_GbE = "Eth_GbE:" + a.get("%s" % test3)
        elif str.find(test3, "Gigabit") != -1:
            print("有线MAC:", a.get("%s" % test3))
            mac1.append(a.get("%s" % test3))
            Eth_GbE = "Eth_GbE:" + a.get("%s" % test3)
        elif str.find(test3, "Wireless") != -1:
            print("无线MAC:", a.get("%s" % test3))
            mac1.append(a.get("%s" % test3))
            Wireless = "Wireless:" + a.get("%s" % test3)
        elif str.find(test3, "无线") != -1:
            print("无线MAC:", a.get("%s" % test3))
            mac1.append(a.get("%s" % test3))
            Wireless = "Wireless:" + a.get("%s" % test3)
        elif str.find(test3, "Centrino") != -1:
            print("无线MAC:", a.get("%s" % test3))
            mac1.append(a.get("%s" % test3))
            Wireless = "Wireless:" + a.get("%s" % test3)


# 字符串切片
# Python replace() 方法把字符串中的 old（旧字符串） 替换成 new(新字符串)，如果指定第三个参数max，则替换不超过 max 次。
Eth_GbE_After = Eth_GbE.replace(":","-")
Wireless_After = Wireless.replace(":","-")

print(Eth_GbE_After[-17:])
print(Wireless_After[-17:])
list_all = [SerialNumber,localtime,"#",name,Eth_GbE_After[-17:],Wireless_After[-17:],"#",hostname]

# print(mac1)
# mac1[0]为有线mac , mac1[1]为无线mac
# print(",".join(mac1))



file1 = os.getcwd()
file2 = '\\'+name+'_infotest.txt'

# file3 = open(file1+file2 ,'w')
# file3.write('当前时间:'+localtime+'\n'+'使用人:'+name+'\n'+'SN:'+SerialNumber+'\n'+'计算机名:'+hostname+'\n'+'本地ip:'+ip+'\n'+'有线mac:'+Eth_GbE[-17:]+'\n''无线mac:'+Wireless[-17:])
# file3.close()

try:

    # 将信息写入excel表格
    wb = load_workbook(file1+'\\'+filename,keep_vba=True)

    # keep_vba=True 需要指定属性keep_vba=True 才能保存xlsm文档
    # wb = load_workbook(file1+'\\test_Excel.xlsm',keep_vba=True)
    # ws = wb.active                #获取当前活跃的sheet,默认是第一个sheetb

    sheets = wb.get_sheet_names()  # 从名称获取sheet
    ws = wb.get_sheet_by_name(sheets[0])
    # print(ws)


    ws.append(list_all)

    rows = ws.rows
    columns = ws.columns
    # 迭代所有的行
    # for row in rows:
    #     line = [col.value for col in row]
    #
    # # 通过坐标读取值
    #
    # cell_11 = ws.cell(row=1, column=1).value

    # 将此属性设置为False（默认），以保存为文档：
    wb.template = False

    # 您可以指定属性template = True，以将工作簿另存为模板：
    # wb.template = True



    wb.save(filename=file1+'\\'+filename)
    # wb.save(filename=file1+'\\test_Excel.xlsm')
    #
except Exception as e:
    print("Exception",e)
    input('Please press enter key to exit ...')

